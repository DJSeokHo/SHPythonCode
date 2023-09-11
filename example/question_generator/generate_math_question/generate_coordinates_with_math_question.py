# _*_ coding: utf-8 _*_
# @Time : 2023/04/07 5:32 PM
# @Author : Coding with cat
# @File : generate_math_question
# @Project : SHPythonCode

import random

import openpyxl

from framewrok.utility.log_utility import ILog


def generate_math_problem(number: int, symbol: str):
    operators = ["+", "-", "*", "/"]
    number_range = range(1, 10)

    # 生成包含已知的正整数A的简单数学题目，结果等于A
    while True:
        # 生成四个随机数
        numbers = [random.choice(number_range) for i in range(4)]
        # 生成三个随机运算符
        op1, op2 = random.choices(operators, k=2)
        # 计算结果
        result = eval(f"{numbers[0]} {op1} {numbers[1]} {op2} {numbers[2]} {operators[3]} {numbers[3]}")
        # 如果结果等于已知的正整数A，则输出数学题目
        if result == number:
            # problem_str = f"{numbers[0]} {op1} {numbers[1]} {op2} {numbers[2]} {operators[3]} {numbers[3]} = {number}"
            problem_str = f"{symbol} = {numbers[0]} {op1} {numbers[1]} {op2} {numbers[2]} {operators[3]} {numbers[3]}"
            problem_str = problem_str.replace("*", "x")
            problem_str = problem_str.replace("/", "÷")
            return problem_str


def replace_coordinates(latitude: float, longitude: float):
    lat_str = str(latitude)
    long_str = str(longitude)

    lat_digits = [int(digit) for digit in lat_str.split(".")[1]]
    long_digits = [int(digit) for digit in long_str.split(".")[1]]

    # select two digits to replace in latitude and longitude respectively
    lat_indices = random.sample(range(len(lat_digits)), 2)
    long_indices = random.sample(range(len(long_digits)), 2)

    number_a = lat_digits[lat_indices[0]]
    number_b = lat_digits[lat_indices[1]]
    number_c = long_digits[long_indices[0]]
    number_d = long_digits[long_indices[1]]

    # replace selected digits with A, B, C, D
    lat_digits[lat_indices[0]] = "A"
    lat_digits[lat_indices[1]] = "B"
    long_digits[long_indices[0]] = "C"
    long_digits[long_indices[1]] = "D"

    # construct replaced latitude and longitude strings
    lat_replaced = lat_str.split(".")[0] + "." + "".join(str(digit) for digit in lat_digits)
    long_replaced = long_str.split(".")[0] + "." + "".join(str(digit) for digit in long_digits)

    # construct dictionary of results
    results = {
        "A": number_a,
        "B": number_b,
        "C": number_c,
        "D": number_d,
        "latitude_replaced": lat_replaced,
        "longitude_replaced": long_replaced
    }

    return results


def read_xlsx_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    coordinates_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None and row[2] is not None:
            coordinates = {
                "address": row[0],
                "latitude": row[1],
                "longitude": row[2]
            }
            coordinates_list.append(coordinates)
    return coordinates_list

def create_math_xlsx_file():
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()

    # 选择默认的工作表
    sheet = wb.active

    # 添加表头
    sheet['A1'] = '员工名字'
    sheet['B1'] = '年龄'
    sheet['C1'] = '工资'

    # 填充数据
    employees = [
        ('张三', 30, 50000),
        ('李四', 25, 45000),
        ('王五', 35, 60000),
        ('赵六', 28, 52000)
    ]

    for row, employee in enumerate(employees, start=2):
        sheet.cell(row=row, column=1, value=employee[0])
        sheet.cell(row=row, column=2, value=employee[1])
        sheet.cell(row=row, column=3, value=employee[2])

    # 保存工作簿到文件
    wb.save('level_1_problem.xlsx')

    # 关闭工作簿
    wb.close()


if __name__ == '__main__':

    # create_math_xlsx_file()

    # dictionary = replace_coordinates(35.19126547, 129.0653389)
    # ILog.debug(__file__, f"{dictionary['latitude_replaced']}, {dictionary['longitude_replaced']}")

    # ILog.debug(__file__, generate_math_problem(number=dictionary["A"], symbol="A"))
    # ILog.debug(__file__, generate_math_problem(number=dictionary["B"], symbol="B"))
    # ILog.debug(__file__, generate_math_problem(number=dictionary["C"], symbol="C"))
    # ILog.debug(__file__, generate_math_problem(number=dictionary["D"], symbol="D"))

    xlsx_file_row_list = read_xlsx_file("수학 문제 장소.xlsx")
    ILog.debug(__file__, len(xlsx_file_row_list))

    problems = []

    for item in xlsx_file_row_list:
        for i in range(0, 10):
            ILog.debug(__file__, item["address"] + " " + str(item["latitude"]) + " " + str(item["longitude"]))
            dictionary = replace_coordinates(item["latitude"], item["longitude"])

            ILog.debug(__file__, f"{dictionary['latitude_replaced']}, {dictionary['longitude_replaced']}")

            ILog.debug(__file__, generate_math_problem(number=dictionary["A"], symbol="A"))
            ILog.debug(__file__, generate_math_problem(number=dictionary["B"], symbol="B"))
            ILog.debug(__file__, generate_math_problem(number=dictionary["C"], symbol="C"))
            ILog.debug(__file__, generate_math_problem(number=dictionary["D"], symbol="D"))

            # ILog.debug(__file__, "=======================================")

            time = "~"
            userId = "admin"
            startLocationName = None
            endLocationName = item["address"]
            endLatitude = item["latitude"]
            endLongitude = item["longitude"]
            questDirection = "다음 문제를 풀고 대답이 제시하는 위치에 이동하여 위치를 인증하세요."
            questContent = (
                    f"\n{dictionary['latitude_replaced']}, {dictionary['longitude_replaced']}\n\n"
                    + generate_math_problem(number=dictionary["A"], symbol="A") + "\n"
                    + generate_math_problem(number=dictionary["B"], symbol="B") + "\n"
                    + generate_math_problem(number=dictionary["C"], symbol="C") + "\n"
                    + generate_math_problem(number=dictionary["D"], symbol="D")
            )
            questType = 1
            answerType = 2
            answerArray = None
            problemContentLink = None
            problemDifficult = 1
            problemTranslate = True
            problemImageLink = None
            problemConfirmed = 0

            problems.append(
                (time, userId, startLocationName, endLocationName, endLatitude, endLongitude,
                 questDirection, questContent, questType, answerType, answerArray,
                 problemContentLink, problemDifficult, problemTranslate, problemImageLink, problemConfirmed)
            )

    ILog.debug(__file__, f"create xlsx file item count {len(problems)}")
    wb = openpyxl.Workbook()
    sheet = wb.active

    # 添加表头
    sheet["A1"] = "time"
    sheet["B1"] = "userId"
    sheet["C1"] = "startLocationName"
    sheet["D1"] = "endLocationName"
    sheet["E1"] = "endLatitude"
    sheet["F1"] = "endLongitude"
    sheet["G1"] = "questDirection"
    sheet["H1"] = "questContent"
    sheet["I1"] = "questType"
    sheet["J1"] = "answerType"
    sheet["K1"] = "answerArray"
    sheet["L1"] = "problemContentLink"
    sheet["M1"] = "problemDifficult"
    sheet["N1"] = "problemTranslate"
    sheet["O1"] = "problemImageLink"
    sheet["P1"] = "problemConfirmed"

    for row, problems in enumerate(problems, start=2):
        sheet.cell(row=row, column=1, value=problems[0])
        sheet.cell(row=row, column=2, value=problems[1])
        sheet.cell(row=row, column=3, value=problems[2])
        sheet.cell(row=row, column=4, value=problems[3])
        sheet.cell(row=row, column=5, value=problems[4])
        sheet.cell(row=row, column=6, value=problems[5])
        sheet.cell(row=row, column=7, value=problems[6])
        sheet.cell(row=row, column=8, value=problems[7])
        sheet.cell(row=row, column=9, value=problems[8])
        sheet.cell(row=row, column=10, value=problems[9])
        sheet.cell(row=row, column=11, value=problems[10])
        sheet.cell(row=row, column=12, value=problems[11])
        sheet.cell(row=row, column=13, value=problems[12])
        sheet.cell(row=row, column=14, value=problems[13])
        sheet.cell(row=row, column=15, value=problems[14])
        sheet.cell(row=row, column=16, value=problems[15])

    # 保存工作簿到文件
    wb.save('level_1_problem.xlsx')

    # 关闭工作簿
    wb.close()


